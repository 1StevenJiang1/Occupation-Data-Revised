from openpyxl import Workbook
from openpyxl.styles import PatternFill
if __name__ == "__main__":
    # Create a new Excel workbook and select the active sheet
    wb = Workbook()
    sheet = wb.active

    # Data to be written to the cells
    data = [
        [1, 2, 3],
        [4, 5, 6],
        [7, 8, 9],
    ]

    # Write the data to the cells
    for row_index, row_data in enumerate(data, start=1):
        for col_index, cell_value in enumerate(row_data, start=1):
            sheet.cell(row=row_index, column=col_index, value=cell_value)

    # Apply cell formatting (highlight) to specific cells
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # For example, let's highlight cells (1,1) and (2,2)
    highlight_cells = [(1, 1), (2, 2)]

    for cell_coords in highlight_cells:
        row, col = cell_coords
        sheet.cell(row=row, column=col).fill = highlight_fill

    # Save the workbook to a file
    wb.save("highlighted_cells.xlsx")
