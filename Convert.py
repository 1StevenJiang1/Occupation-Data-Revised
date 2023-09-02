from typing import TextIO
from typing import List, Dict

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill


class ExcelHandler:
    sheet_names: List[str]
    column_names: List[str]
    columns: Dict[str, List[str]]

    def __init__(self, sheet_names: List[str], column_names: List[str], columns: Dict[str, List[str]]):
        self.sheet_names = sheet_names
        self.column_names = column_names
        self.columns = columns

    def to_excel(self, filename: str) -> None:
        df = pd.DataFrame(self.columns)
        df.to_excel(filename, index=False)


def to_list(file: TextIO) -> List[str]:
    f = file.readlines()
    for i in range(len(f)):
        f[i] = f[i].strip()
        f[i] = f[i].split('\t')
    return f


def get_soc_code(df) -> List[str]:
    df.fillna('', inplace=True)
    df.iloc[:, 0] = df.iloc[:, 0].str.strip()
    column_1_list = df.iloc[:627, 0].tolist()
    print(column_1_list)
    return column_1_list


def remove_nan(lst: List[str]) -> None:
    for i in range(len(lst) - 1, -1, -1):
        if lst[i] == '':
            lst.pop(i)
    print(lst)


def add_decimal(lst: List[str]) -> None:
    for i in range(len(lst)):
        if lst[i] != 'none':
            lst[i] = lst[i] + '.00'
    print(lst)


def add_new_column(path: str, lst: List[str]) -> None:

    # Load the Excel file
    file_path = path
    workbook = openpyxl.load_workbook(file_path)

    # Select the first sheet (you can also specify the sheet name)
    sheet = workbook.active

    # Replace the first column with the elements from the list 'lst'
    for row_index, value in enumerate(lst, start=1):
        sheet.cell(row=row_index, column=1, value=value)

    # Save the modified workbook to a new file
    output_file_path = 'output_version.xlsx'
    workbook.save(output_file_path)


def highlight_cells_not_in_other_column(path, sheet_name):
    # Load the Excel file
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]

    # Get the "new" and "O*NET-SOC Code" columns
    new_column = list(sheet.columns)[0]
    soc_code_column = list(sheet.columns)[1]

    # Set up the highlight fill color (e.g., yellow)
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00",
                                 fill_type="solid")

    # Iterate through the "new" column
    for cell in new_column[1:]:  # Skip the header row
        if cell.value not in [soc_cell.value for soc_cell in
                              soc_code_column[1:]]:
            cell.fill = highlight_fill

    # Save the modified workbook to a new file
    output_file_path = 'output.xlsx'
    workbook.save(output_file_path)


def read_from_excel(file_name: str, sheet_name: str, column_letter: str) -> List[str]:
    """write the column into the file specified by file name"""
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(file_name)

    # Choose the specific sheet
    worksheet = workbook[sheet_name]  # Replace 'Sheet1' with the actual sheet name

    # Choose the column you want to convert into a list (e.g., Column A)

    # Get the values from the selected column using a list comprehension
    column_values = []
    for cell in worksheet[column_letter]:
        column_values.append(cell.value)

    # Print the list of values
    # print(column_values)
    return column_values


def strip_super_script(lsts: List[List[str]]) -> List[List[str]]:
    """Return a list of list of str. If the element in the sublist is a number and has super script,
    remove it. The element is left unchanged otherwise.
    >>> a = [['Canada', '101'], [None, '20']]
    >>> result = strip_super_script(a)
    >>> print(result)
    >>> [['Canada', '10'], [None, '20']]
    """
    for lst in lsts:
        for i in range(len(lst)):
            if lst[i]:
                lst[i] = str(lst[i])
    print(lsts)
    for lst in lsts:
        for i in range(len(lst)):
            # first check whether the item is None, and then see whether
            # it is a number
            if lst[i] and lst[i][-1].isdigit():
                if 0 < int(lst[i][-1]) < 5:
                    lst[i] = lst[i][:-1]
    return lsts


def deal_with_merged_cells(lsts: List[List[str]]) -> List[List[str]]:
    """Return a list of list of str, where the resulting None created by merged cells are filled
    by corresponding values
    >>> b = [['Canada', '10'], [None, '20'], [None, None]]
    >>> b = deal_with_merged_cells(b)
    >>> print(b)
    >>> [['Canada', '10'], ['Canada', '20'], ['Canada', '20']]
    """
    for lst in lsts:
        i = 0
        # This is primarily because the data is incorrectly formatted, resulting a lot of empty cells
        # at the beginning of each column. This while loop is to skip these blank cells and get to the
        # first cell that actually contains value
        while not lst[i]:
            i += 1
        cell_value = None
        while i < len(lst):
            # if the cell contains a value, update cell_value. If the cell contains None, then it should
            # inherit value from its predecessor in the column.
            if lst[i]:
                cell_value = lst[i]
            if not lst[i]:
                lst[i] = cell_value
            i += 1
    return lsts


def fill_zero(lsts: List[List[str]]) -> List[List[str]]:
    """Return a list of list where each element, if the element is a number, is a four-digit number. Use 0
    as a place-holder if the number does not have the hundredth and thousandth digit.
    >>> c = [['Canada', '10'], ['Canada', '20']]
    >>> c = fill_zero(c)
    >>> print(c)
    >>> [['Canada', '0010'], ['Canada', '0020']]]
    """
    for lst in lsts:
        i = 0
        # This is primarily because the data is incorrectly formatted, resulting a lot of empty cells
        # at the beginning of each column. This while loop is to skip these blank cells and get to the
        # first cell that actually contains value
        while not lst[i]:
            i += 1
        while i < len(lst):
            if lst[i].isdigit():
                missed_0 = 4 - len(lst[i])
                lst[i] = missed_0 * '0' + lst[i]
            i += 1
    return lsts


def clean_occ_crosswalk(f_path: str, sheet_name: str, column_names: List[str], output_name: str) -> \
        List[List[str]]:
    """With the Excel file on the path f_path and information on sheet sheet_name and columns specified by
    column_name, return the cleaned version of it. The function performs a sequence of operations, where it first
    strips superscripts, then handle the merged cells, and lastly fill each cell with appropriate number of zeros.
    For details of each helper function, see sample usage above.
    """
    input_lsts = []
    input_dict = {}
    # Read input from the file
    for column_name in column_names:
        input_lst = read_from_excel(f_path, sheet_name, column_name)
        input_lsts.append(input_lst)

    # Clean data
    input_lsts = strip_super_script(input_lsts)
    input_lsts = deal_with_merged_cells(input_lsts)
    input_lsts = fill_zero(input_lsts)

    # Write into a new Excel file
    i = 0
    for column_name in column_names:
        input_dict[column_name] = input_lsts[i]
    handler = ExcelHandler(['Sheet1'], column_names, input_dict)
    handler.to_excel(output_name)
    return input_lsts


def clean_onet(lsts: List[List[str]]) -> List[List[str]]:
    """Return a list of list of str where each o*net occupation code is correctly associated an
    occupation name that has values
    """
    temp = [[], [], []]
    for i in range(len(lsts[1])):
        if lsts[1][i] in lsts[0]:
            temp[0].append(lsts[1][i])
            temp[1].append(lsts[2][i])
            temp[2].append(lsts[3][i])
    return temp


def set_up_for_merge(lsts: List[List[str]]) -> List[List[str]]:
    """Return a list of list of str representing a spreadsheet. The function first strip off the decimal
    places in each code. Then it removes all duplicated values resulted from the previous operation. If a code
    has multiple occurrences in the sheet, only the first one is kept."""
    if len(lsts) != 3:
        raise ValueError("The number of input column is incorrect. Please check again")

    for i in range(1, len(lsts[0])):
        lsts[0][i] = lsts[0][i][:7]
    seen = None
    temp = [[], [], []]
    for i in range(len(lsts[0])):
        item = lsts[0][i]
        if lsts[0][i] != seen:
            temp[0].append(lsts[0][i])
            temp[1].append(lsts[1][i])
            temp[2].append(lsts[2][i])
            seen = item
    return temp


def merge(lsts: List[List[str]]) -> List[List[str]]:
    """Return a list of list of str representing a spreadsheet. The first column contains IPUMS code, where the
    following cells in the same rows are the corresponding O*Net data. Note that some codes are in one file but
    not in the other. These are omitted in the final output. """
    # the First column will be IPUMS Code, the second one IPUMS Title, the third one O*Net Code, and the last one
    # O*Net Title
    temp = [[], [], [], []]
    for i in range(len(lsts[2])):
        item = lsts[0][i]
        if item in lsts[2]:
            onet_position = lsts[2].index(item)
            temp[0].append(item)
            temp[1].append(lsts[1][i])
            temp[2].append(lsts[2][onet_position])
            temp[3].append(lsts[3][onet_position])
    print(temp)

    return temp

def toy():
    return None

if __name__ == "__main__":
    lst1 = read_from_excel("Occupation Data (dup).xlsx", "Sheet1", 'A')
    lst2 = read_from_excel("Occupation Data (dup).xlsx", 'Sheet1', 'B')
    lst3 = read_from_excel("Occupation Data (dup).xlsx", 'Sheet1', 'C')
    lst4 = read_from_excel("Occupation Data (dup).xlsx", 'Sheet1', 'D')
    columns = [lst1, lst2, lst3, lst4]
    a = clean_onet(columns)
    a = set_up_for_merge(a[:3])
    lst5 = read_from_excel("2010-occ-codes-with-crosswalk-from-2002-2011.xlsx", '2010OccCodeList', 'D')
    lst6 = read_from_excel("2010-occ-codes-with-crosswalk-from-2002-2011.xlsx", '2010OccCodeList', 'B')
    lst5 = lst5 + [np.nan] * (len(a[0]) - len(lst5))
    lst6 = lst6 + [np.nan] * (len(a[0]) - len(lst6))
    a.insert(0, lst6)
    a.insert(0, lst5)
    a = merge(a)
    new_dict = {'IPUMS Code': a[0], 'IPUMS Data': a[1], 'O*Net Code': a[2], 'Title': a[3]}
    handler = ExcelHandler(['Sheet1'], ['O*Net Code', 'Title', 'Description'], new_dict)
    handler.to_excel("Occupation Data (merged1).xlsx")

    # print(lst1)
    # print(lst2)
    # for item in lst2:
    #     if item in lst1:
    #         tem.append(item)
    # print(tem)


    # lst1 = read_from_excel("2010-occ-codes-with-crosswalk-from-2002-2011.xlsx", "2010OccCodeList", 'H')
    # lst2 = read_from_excel("2010-occ-codes-with-crosswalk-from-2002-2011.xlsx", '2010OccCodeList', 'I')
    # lst3 = read_from_excel("2010-occ-codes-with-crosswalk-from-2002-2011.xlsx", '2010OccCodeList', 'J')
    # lst4 = read_from_excel("2010-occ-codes-with-crosswalk-from-2002-2011.xlsx", '2010OccCodeList', 'K')
    # lst5 = read_from_excel("2010-occ-codes-with-crosswalk-from-2002-2011.xlsx", '2010OccCodeList', 'L')
    # columns = [lst1, lst2, lst3, lst4, lst5]
    # a = strip_super_script(columns)
    # print("Result of strip_super_script")
    # print(a)
    # a = deal_with_merged_cells(a)
    # print("Result of deal_with_merged_cell")
    # print(a)
    # a = fill_zero(a)
    # print("Result of fill_zero")
    # print(a)
    # new_dict = {'A': a[0], 'B': a[1], 'C': a[2], 'D': a[3], 'E': a[4]}
    # handler1 = ExcelHandler(['Sheet1'], ['A', 'B', 'C', 'D', 'E'], new_dict)
    # handler1.to_excel('clean_coding.xlsx')

